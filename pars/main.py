

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from pars import dict_products


wb = openpyxl.Workbook()
wb.guess_types = True
ws = wb.active

sch = 0
count_g = 1
count_v = 0
low_price = []
high_price = []
middle_price = []
for i in dict_products.values():
    d = []
    for j in i:
        if len(j) != 0 and j[0].isdigit():
            if j[-1].isdigit():
                d.append(int(j))
                if sch > 0:
                    if int(i[1]) > (int(j)+13):
                        low_price.append([count_g, count_v])
                    if int(i[1]) < (int(j)-13):
                        high_price.append([count_g, count_v])
                    if (int(j)-13) <= int(i[1]) <= (int(j)+13):
                        middle_price.append([count_g, count_v])

            else:
                d.append(j)
                j = j.split()[0]
                if sch > 0:
                    if int(i[1]) > (int(j)+13):
                        low_price.append([count_g, count_v])
                    if int(i[1]) < (int(j)-13):
                        high_price.append([count_g, count_v])
                    if (int(j)-13) <= int(i[1]) <= (int(j)+13):
                        middle_price.append([count_g, count_v])

        else:
            d.append(j)
        count_v += 1
    count_v = 0
    count_g += 1
    sch += 1
    ws.append(d)



ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 25
ws.column_dimensions["F"].width = 25

for i in range(1, ws.max_row+1):
    ws.row_dimensions[i].height = 25

for i in ws[1]:
    i.alignment = Alignment(horizontal="center", vertical="center")
    i.font = Font(size=12, bold=True)

thins = Side(border_style="medium", color="101104")


for i in ws:
    for j in i:
        j.number_format = '# ##0'
        j.border = Border(top=thins, bottom=thins, left=thins, right=thins)


for i in low_price:
    ws[i[0]][i[1]].fill = PatternFill('solid', fgColor="F34153")
    ws[i[0]][i[1]].alignment = Alignment(horizontal="center", vertical="center")

for i in high_price:
    ws[i[0]][i[1]].fill = PatternFill('solid', fgColor="21D214")
    ws[i[0]][i[1]].alignment = Alignment(horizontal="center", vertical="center")

for i in middle_price:
    ws[i[0]][i[1]].fill = PatternFill('solid', fgColor="E97D11")
    ws[i[0]][i[1]].alignment = Alignment(horizontal="center", vertical="center")

for i in range(1, len(dict_products)+1):
    if i > 1:
        ws[i][1].fill = PatternFill('solid', fgColor="EAF728")
        ws[i][1].alignment = Alignment(horizontal="center", vertical="center")
        ws[i][0].alignment = Alignment(vertical="center")
        ws[i][0].font = Font(size=12)

wb.save('products.xlsx')
wb.close()


